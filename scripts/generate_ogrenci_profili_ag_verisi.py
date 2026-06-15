import hashlib
import json
import math
import re
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[1]
SOURCE_MARKDOWN = Path(r"C:\Users\vedat\Downloads\hamveriler.md")
ORTAOGRETIM_SOURCE_DIR = Path("C:/Users/vedat/Desktop/orta\u00f6\u011fretim")
DIN_OGRETIMI_SOURCE = Path(r"C:\Users\vedat\Desktop\din kültürü.xlsx")
SAGLIK_ORTAOGRETIM_SOURCE = Path(r"C:\Users\vedat\Desktop\ortaöğretim\Öğrenci Profili_sağlıkvetrafik.xlsx")
OUTPUT_PATH = REPO_ROOT / "PersonelTakipSistemi" / "wwwroot" / "data" / "ogrenci_profili_ag_verisi.json"

TEMEL_EGITIM = "Temel Eğitim"
ORTAOGRETIM = "Ortaöğretim"
DIN_OGRETIMI = "Din Öğretimi"
EXCLUDED_LESSONS = {"Trafik Güvenliği"}
EXCLUDED_PROFILES = {"Duyarlı", "Mütevazı"}

ORTAOGRETIM_SOURCE_SHEETS = {
    "Biyoloji.xlsx": ("Fen Bilimleri",),
    "Co\u011frafya.xlsx": ("Co\u011frafya",),
    "Felsefe.xlsx": ("Felsefe",),
    "Fizik.xlsx": ("Fizik",),
    "Kimya.xlsx": ("Fen Bilimleri",),
    "Matematik.xlsx": ("Lise Matematik",),
    "Tarih ve TC \u0130nk\u0131lap.xlsx": ("Tarih", "T.C. \u0130nk\u0131lap Tarihi"),
    "T\u00fcrk Dili ve Edebiyat\u0131.xlsx": ("T\u00fcrk Dili ve Edebiyat\u0131",),
}

ORTAOGRETIM_HISTORY_SOURCE_FILE = "Tarih ve TC \u0130nk\u0131lap.xlsx"
ORTAOGRETIM_HISTORY_LESSON = "Tarih ve T.C. \u0130nk\u0131lap"

PROFILE_ALIASES = {
    "ahlakli": "Ahlaklı",
    "bilge": "Bilge",
    "cesaret": "Cesaretli",
    "cesaretli": "Cesaretli",
    "cesur": "Cesaretli",
    "estetik": "Estetik",
    "estetiik": "Estetik",
    "iradeli": "İradeli",
    "mehametli": "Merhametli",
    "merhametli": "Merhametli",
    "saglik": "Sağlıklı",
    "saglikli": "Sağlıklı",
    "sorulayici": "Sorgulayıcı",
    "sorgulayici": "Sorgulayıcı",
    "uretken": "Üretken",
    "uretkenken": "Üretken",
    "uretkenlik": "Üretken",
    "vatansever": "Vatansever",
    "vatanseverlik": "Vatansever",
}

PROFILE_RING_ORDER = [
    "Vatansever",
    "Ahlaklı",
    "Bilge",
    "Cesaretli",
    "Estetik",
    "İradeli",
    "Merhametli",
    "Sağlıklı",
    "Sorgulayıcı",
    "Üretken",
]

PROFILE_COLORS = {
    "Vatansever": "#a45d72",
    "Ahlaklı": "#c46283",
    "Bilge": "#a45d92",
    "Cesaretli": "#8263a8",
    "Estetik": "#6f77b5",
    "İradeli": "#36899b",
    "Merhametli": "#47a89b",
    "Sağlıklı": "#66b98e",
    "Sorgulayıcı": "#86ad71",
    "Üretken": "#d6955f",
}

LESSON_ALIASES = {
    "biyoloji": "Biyoloji",
    "cografya": "Coğrafya",
    "din kulturu ve ahlak bilgisi": "Din Kültürü ve Ahlak Bilgisi",
    "felsefe": "Felsefe",
    "fen bilimleri": "Fen Bilimleri",
    "fizik": "Fizik",
    "hayat bilgisi": "Hayat Bilgisi",
    "kimya": "Kimya",
    "kur an i kerim dersi": "Kur’an-ı Kerim Dersi",
    "matematik": "Matematik",
    "muzik": "Müzik",
    "peygamberimizin hayati": "Peygamberimizin Hayatı",
    "sosyal bilgiler": "Sosyal Bilgiler",
    "t c inkilap tarihi ve ataturkculuk": "T.C. İnkılap Tarihi ve Atatürkçülük",
    "tarih": "Tarih",
    "temel dini bilgiler": "Temel Dinî Bilgiler",
    "turk dili ve edebiyati": "Türk Dili ve Edebiyatı",
    "turkce": "Türkçe",
}

PALETTE = {
    "MERKEZ": "#39424e",
    "PROFIL": "#8E5A85",
    "KATEGORI": "#546E7A",
    "DERS": "#455e75",
    "SINIF": "#607D8B",
    "UNITE_TEMA": "#7A9E9F",
}

HC_PALETTE = {
    "MERKEZ": "#ffffff",
    "PROFIL": "#ff80ab",
    "KATEGORI": "#ffffff",
    "DERS": "#ffea00",
    "SINIF": "#ffffff",
    "UNITE_TEMA": "#00e5ff",
}

GROUP_ORDER = ["MERKEZ", "KATEGORI", "PROFIL", "DERS", "SINIF", "UNITE_TEMA"]


def ascii_key(value):
    replacements = str.maketrans(
        {
            "ı": "i",
            "İ": "i",
            "ğ": "g",
            "Ğ": "g",
            "ü": "u",
            "Ü": "u",
            "ş": "s",
            "Ş": "s",
            "ö": "o",
            "Ö": "o",
            "ç": "c",
            "Ç": "c",
        }
    )
    value = (value or "").translate(replacements)
    value = unicodedata.normalize("NFKD", value)
    value = "".join(ch for ch in value if not unicodedata.combining(ch))
    return value.casefold()


def clean(value):
    value = str(value or "").replace("<br>", ",").replace("\n", " ").replace("\r", " ")
    value = re.sub(r"\s+", " ", value)
    return value.strip(" \t;,/|")


def slug(value, prefix):
    key = ascii_key(value)
    key = re.sub(r"[^a-z0-9]+", "_", key).strip("_")
    if len(key) > 70:
        digest = hashlib.sha1(value.encode("utf-8")).hexdigest()[:8]
        key = f"{key[:58]}_{digest}"
    return f"{prefix}_{key or 'bos'}".upper()


def normalize_class(value):
    value = clean(value)
    if "hazirlik" in ascii_key(value):
        return "Hazırlık Sınıfı"
    range_match = re.search(r"(\d+)\s*[-–]\s*(\d+)", value or "")
    if range_match:
        return f"{int(range_match.group(1))}-{int(range_match.group(2))}. Sınıf"
    match = re.search(r"(\d+)", value or "")
    return f"{int(match.group(1))}. Sınıf" if match else value


def normalize_lesson(value):
    value = clean(value)
    key = re.sub(r"[^a-z0-9]+", " ", ascii_key(value)).strip()
    if key in LESSON_ALIASES:
        return LESSON_ALIASES[key]
    value = value.title().replace("\u0307", "")
    return value.replace("T.C. Inkılap", "T.C. İnkılap")


def apply_curriculum_rules(record):
    if (
        record.get("kategori") == ORTAOGRETIM
        and record.get("file") == ORTAOGRETIM_HISTORY_SOURCE_FILE
    ):
        record["ders"] = ORTAOGRETIM_HISTORY_LESSON
    elif record.get("ders") == "Matematik":
        if record.get("kategori") == TEMEL_EGITIM:
            record["ders"] = "Matematik (Temel Eğitim)"
        elif record.get("kategori") == ORTAOGRETIM:
            record["ders"] = "Matematik (Ortaöğretim)"
    return record


def normalize_unit(value):
    value = clean(value)
    value = re.sub(r"\s*/\s*$", "", value)
    return value


def unit_key(value):
    key = ascii_key(value)
    key = re.sub(r"[^a-z0-9]+", " ", key)
    return re.sub(r"\s+", " ", key).strip()


def display_score(value):
    letters = [ch for ch in value if ch.isalpha()]
    lowercase = sum(1 for ch in letters if ch.islower())
    return lowercase, -len(value)


def normalize_profile_token(value):
    token = clean(value)
    key = re.sub(r"[^a-z0-9]+", "", ascii_key(token))
    if not key:
        return ""
    profile = PROFILE_ALIASES.get(key, token[:1].upper() + token[1:])
    return "" if profile in EXCLUDED_PROFILES else profile


def split_profiles(value):
    value = clean(value)
    if not value:
        return []

    parts = re.split(r"[,;/]+", value)
    result = []
    known_pattern = re.compile(
        r"\b("
        + "|".join(
            sorted(
                (
                    "Ahlaklı",
                    "Bilge",
                    "Cesaret",
                    "Cesaretli",
                    "Estetik",
                    "İradeli",
                    "Mehametli",
                    "Merhametli",
                    "Sağlık",
                    "Sağlıklı",
                    "Sorgulayıcı",
                    "Uretken",
                    "Üretken",
                    "Üretkenlik",
                    "Vatansever",
                    "Vatanseverlik",
                    "ahlaklı",
                    "bilge",
                    "cesaret",
                    "iradeli",
                    "merhametli",
                    "sağlıklı",
                    "sorgulayıcı",
                    "vatansever",
                    "üretken",
                    "üretkenlik",
                ),
                key=len,
                reverse=True,
            )
        )
        + r")\b",
        flags=re.IGNORECASE,
    )

    for part in parts:
        matches = known_pattern.findall(part)
        tokens = matches if len(matches) > 1 else [part]
        for token in tokens:
            profile = normalize_profile_token(token)
            if profile and profile not in result:
                result.append(profile)
    return result


def scrub_profile_raw(value):
    value = clean(value)
    if not value:
        return ""

    value_key = ascii_key(value)
    excluded_keys = {ascii_key(profile) for profile in EXCLUDED_PROFILES}
    if not any(key in value_key for key in excluded_keys):
        return value

    return "; ".join(split_profiles(value))


def split_markdown_row(line):
    return [cell.strip() for cell in line.strip().strip("|").split("|")]


def read_records():
    text = SOURCE_MARKDOWN.read_text(encoding="utf-8")
    records = []
    skipped = []
    current_file = ""
    current_class = ""
    current_lesson = ""

    for line_number, line in enumerate(text.splitlines(), start=1):
        line = line.rstrip()
        if line.startswith("## ") and not line.startswith("## Genel"):
            current_file = line[3:].strip()
            current_class = ""
            current_lesson = ""
            continue

        if line.startswith("### ") and "/" in line:
            heading = line[4:].strip()
            raw_class, raw_lesson = [part.strip() for part in heading.split("/", 1)]
            current_class = normalize_class(raw_class)
            current_lesson = normalize_lesson(raw_lesson)
            continue

        if not line.startswith("|") or line.startswith("|---") or "Excel Satırı" in line:
            continue

        cells = split_markdown_row(line)
        if len(cells) < 4 or not re.fullmatch(r"\d+", cells[0] or ""):
            continue

        excel_row = int(cells[0])
        unit = normalize_unit(cells[1])
        raw_main = clean(cells[2])
        raw_support = clean(cells[3])
        main_profiles = split_profiles(raw_main)
        support_profiles = split_profiles(raw_support)

        if not current_file or not current_class or not current_lesson or not unit or not main_profiles:
            skipped.append(
                {
                    "markdownLine": line_number,
                    "file": current_file,
                    "row": excel_row,
                    "kategori": TEMEL_EGITIM,
                    "sinif": current_class,
                    "ders": current_lesson,
                    "unite": unit,
                    "anaRaw": raw_main,
                    "destekRaw": raw_support,
                }
            )
            continue

        records.append(
            {
                "markdownLine": line_number,
                "file": current_file,
                "row": excel_row,
                "kategori": TEMEL_EGITIM,
                "sinif": current_class,
                "ders": current_lesson,
                "unite": unit,
                "anaProfilRaw": scrub_profile_raw(raw_main),
                "destekProfilRaw": scrub_profile_raw(raw_support),
                "anaProfiller": main_profiles,
                "destekProfiller": support_profiles,
            }
        )

    unit_labels = {}
    for record in records:
        key = unit_key(record["unite"])
        if key not in unit_labels or display_score(record["unite"]) > display_score(unit_labels[key]):
            unit_labels[key] = record["unite"]

    for record in records:
        record["unite"] = unit_labels[unit_key(record["unite"])]

    return records, skipped


def find_header_indexes(row):
    headers = {
        re.sub(r"[^a-z0-9]+", " ", ascii_key(clean(value))).strip(): index
        for index, value in enumerate(row)
    }
    return {
        "sinif": headers.get("sinif seviyesi"),
        "ders": headers.get("ders adi") or headers.get("ders"),
        "unite": headers.get("unite ogrenme alani tema") or headers.get("tema"),
        "ana": headers.get("ana profil"),
        "destek": headers.get("destekleyici profil"),
    }


def iter_sheet_records(path, sheet_name, skipped, kategori=ORTAOGRETIM):
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    header_indexes = None
    header_row = 0
    for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        values = [clean(value) for value in row]
        if any("ana profil" == ascii_key(value) for value in values):
            header_indexes = find_header_indexes(values)
            header_row = row_number
            break

    required = ("sinif", "ders", "unite", "ana", "destek")
    if not header_indexes or any(header_indexes[key] is None for key in required):
        skipped.append(
            {
                "file": path.name,
                "sheet": sheet_name,
                "row": header_row,
                "kategori": kategori,
                "reason": "Başlık satırı okunamadı.",
            }
        )
        return []

    records = []
    current_class = ""
    current_lesson = ""
    for row_number, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        values = [clean(value) for value in row]
        raw_class = values[header_indexes["sinif"]] if header_indexes["sinif"] < len(values) else ""
        raw_lesson = values[header_indexes["ders"]] if header_indexes["ders"] < len(values) else ""
        raw_unit = values[header_indexes["unite"]] if header_indexes["unite"] < len(values) else ""
        raw_main = values[header_indexes["ana"]] if header_indexes["ana"] < len(values) else ""
        raw_support = values[header_indexes["destek"]] if header_indexes["destek"] < len(values) else ""

        if not any((raw_class, raw_lesson, raw_unit, raw_main, raw_support)):
            continue

        if raw_class:
            current_class = normalize_class(raw_class)
        if raw_lesson:
            current_lesson = normalize_lesson(raw_lesson)

        if ascii_key(current_lesson) == "fen bilimleri":
            continue

        unit = normalize_unit(raw_unit)
        main_profiles = split_profiles(raw_main)
        support_profiles = split_profiles(raw_support)

        if not current_class or not current_lesson or not unit or not main_profiles:
            skipped.append(
                {
                    "file": path.name,
                    "sheet": sheet_name,
                    "row": row_number,
                    "kategori": kategori,
                    "sinif": current_class,
                    "ders": current_lesson,
                    "unite": unit,
                    "anaRaw": raw_main,
                    "destekRaw": raw_support,
                }
            )
            continue

        records.append(
            {
                "file": path.name,
                "sheet": sheet_name,
                "row": row_number,
                "kategori": kategori,
                "sinif": current_class,
                "ders": current_lesson,
                "unite": unit,
                "anaProfilRaw": scrub_profile_raw(raw_main),
                "destekProfilRaw": scrub_profile_raw(raw_support),
                "anaProfiller": main_profiles,
                "destekProfiller": support_profiles,
            }
        )

    return records


def read_ortaogretim_records():
    records = []
    skipped = []
    for file_name, sheet_names in ORTAOGRETIM_SOURCE_SHEETS.items():
        path = ORTAOGRETIM_SOURCE_DIR / file_name
        if not path.exists():
            skipped.append(
                {
                    "file": file_name,
                    "kategori": ORTAOGRETIM,
                    "reason": "Excel dosyası bulunamadı.",
                }
            )
            continue
        for sheet_name in sheet_names:
            records.extend(iter_sheet_records(path, sheet_name, skipped))
    return records, skipped


def read_din_ogretimi_records():
    records = []
    skipped = []
    if not DIN_OGRETIMI_SOURCE.exists():
        return records, [
            {
                "file": DIN_OGRETIMI_SOURCE.name,
                "kategori": DIN_OGRETIMI,
                "reason": "Din Öğretimi Excel dosyası bulunamadı.",
            }
        ]

    workbook = load_workbook(DIN_OGRETIMI_SOURCE, read_only=True, data_only=True)
    for sheet_name in workbook.sheetnames:
        records.extend(iter_sheet_records(DIN_OGRETIMI_SOURCE, sheet_name, skipped, kategori=DIN_OGRETIMI))
    return records, skipped


def make_spreadsheet_record(path, sheet_name, row_number, kategori, sinif, ders, unite, raw_main, raw_support):
    return {
        "file": path.name,
        "sheet": sheet_name,
        "row": row_number,
        "kategori": kategori,
        "sinif": normalize_class(sinif),
        "ders": normalize_lesson(ders),
        "unite": normalize_unit(unite),
        "anaProfilRaw": scrub_profile_raw(raw_main),
        "destekProfilRaw": scrub_profile_raw(raw_support),
        "anaProfiller": split_profiles(raw_main),
        "destekProfiller": split_profiles(raw_support),
    }


def read_ortaogretim_saglik_records():
    records = []
    skipped = []
    if not SAGLIK_ORTAOGRETIM_SOURCE.exists():
        return records, [
            {
                "file": SAGLIK_ORTAOGRETIM_SOURCE.name,
                "kategori": ORTAOGRETIM,
                "reason": "Sağlık Bilgisi Excel dosyası bulunamadı.",
            }
        ]

    workbook = load_workbook(SAGLIK_ORTAOGRETIM_SOURCE, read_only=True, data_only=True)
    sheet_name = "Sağlık ve Trafik Kültürü"
    sheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook[workbook.sheetnames[-1]]
    header_indexes = None
    header_row = 0
    for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        values = [clean(value) for value in row]
        if any("ana profil" == ascii_key(value) for value in values):
            header_indexes = find_header_indexes(values)
            header_row = row_number
            break

    required = ("sinif", "unite", "ana", "destek")
    if not header_indexes or any(header_indexes[key] is None for key in required):
        return records, [
            {
                "file": SAGLIK_ORTAOGRETIM_SOURCE.name,
                "sheet": sheet.title,
                "row": header_row,
                "kategori": ORTAOGRETIM,
                "reason": "Sağlık Bilgisi başlık satırı okunamadı.",
            }
        ]

    for row_number, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        values = [clean(value) for value in row]
        raw_unit = values[header_indexes["unite"]] if header_indexes["unite"] < len(values) else ""
        if not raw_unit or "trafik" in ascii_key(raw_unit):
            continue
        raw_class = values[header_indexes["sinif"]] if header_indexes["sinif"] < len(values) else ""
        raw_main = values[header_indexes["ana"]] if header_indexes["ana"] < len(values) else ""
        raw_support = values[header_indexes["destek"]] if header_indexes["destek"] < len(values) else ""
        record = make_spreadsheet_record(
            SAGLIK_ORTAOGRETIM_SOURCE,
            sheet.title,
            row_number,
            ORTAOGRETIM,
            raw_class,
            "Sağlık Bilgisi",
            raw_unit,
            raw_main,
            raw_support,
        )
        if not record["sinif"] or not record["unite"] or not record["anaProfiller"]:
            skipped.append(
                {
                    "file": SAGLIK_ORTAOGRETIM_SOURCE.name,
                    "sheet": sheet.title,
                    "row": row_number,
                    "kategori": ORTAOGRETIM,
                    "sinif": record["sinif"],
                    "ders": record["ders"],
                    "unite": record["unite"],
                    "anaRaw": raw_main,
                    "destekRaw": raw_support,
                }
            )
            continue
        records.append(record)
    return records, skipped


def count_values(records, field):
    counter = Counter()
    for record in records:
        value = record[field]
        if isinstance(value, list):
            counter.update(value)
        else:
            counter[value] += 1
    return counter


def node_size(count, min_count, max_count, low=16, high=42):
    if max_count == min_count:
        return round((low + high) / 2, 1)
    return round(low + ((count - min_count) / (max_count - min_count)) * (high - low), 1)


def wrap_label(value, max_len=24, max_lines=3):
    words = clean(value).split()
    lines = []
    current = ""
    for word in words:
        candidate = f"{current} {word}".strip()
        if current and len(candidate) > max_len:
            lines.append(current)
            current = word
        else:
            current = candidate
    if current:
        lines.append(current)
    if len(lines) > max_lines:
        lines = lines[:max_lines]
        lines[-1] = lines[-1].rstrip(".") + "..."
    return "\n".join(lines) or clean(value)


def make_node(node_id, name, group, node_type, count, min_count, max_count, shape, **extra):
    border = extra.pop("border", PALETTE[group])
    label_count = "kayıt" if node_type != "profil" else "görünüm"
    node = {
        "id": node_id,
        "name": name,
        "label": f"{wrap_label(name)}\n{count} {label_count}",
        "group": group,
        "nodeType": node_type,
        "shape": shape,
        "size": node_size(count, min_count, max_count),
        "color": {"background": "#ffffff", "border": border},
        "font": {"color": "#111", "multi": True},
        "recordCount": count,
        "title": f"{name}: {count} {label_count}",
    }
    node.update(extra)
    return node


def add_edge(edges, from_id, to_id, relation, refs, dashed=False):
    if from_id == to_id:
        return
    key = (from_id, to_id, relation)
    edge = edges.setdefault(
        key,
        {
            "id": slug(f"{from_id}_{to_id}_{relation}", "E"),
            "from": from_id,
            "to": to_id,
            "arrows": "to",
            "label": relation,
            "relation": relation,
            "value": 0,
            "title": "",
            "font": {"align": "middle", "size": 10},
            "color": {"opacity": 0.48},
            "dashes": dashed,
            "recordRefs": [],
        },
    )
    edge["value"] += len(refs)
    edge["recordRefs"].extend(refs)
    edge["title"] = f"{relation}: {edge['value']} kayıt"


def profile_sort_key(item):
    value, count = item
    if value in PROFILE_RING_ORDER:
        return (0, PROFILE_RING_ORDER.index(value))
    return (1, -count, value)


def sector_point(angle, radius, offset=0):
    return {
        "x": round(math.cos(angle) * radius + offset, 2),
        "y": round(math.sin(angle) * radius, 2),
    }


def distribute_positions(values, angle_lookup, dominant_lookup, radius, spread=0.34):
    buckets = defaultdict(list)
    for value in values:
        buckets[dominant_lookup.get(value, "")].append(value)

    positions = {}
    for dominant, bucket in buckets.items():
        base_angle = angle_lookup.get(dominant, -math.pi / 2)
        total = len(bucket)
        for index, value in enumerate(sorted(bucket)):
            delta = 0 if total == 1 else ((index / (total - 1)) - 0.5) * spread
            positions[value] = sector_point(base_angle + delta, radius + (index % 5) * 42)
    return positions


def category_sort_key(value):
    order = {ORTAOGRETIM: 0, TEMEL_EGITIM: 1, DIN_OGRETIMI: 2}
    return order.get(value, 99), value


def class_sort_key(value):
    if "hazirlik" in ascii_key(value):
        return 0, 0, value
    match = re.search(r"(\d+)", value or "")
    return 1, int(match.group(1)) if match else 999, value


def record_ref(record):
    ref = {
        "file": record["file"],
        "row": record["row"],
        "kategori": record.get("kategori", TEMEL_EGITIM),
    }
    if record.get("sheet"):
        ref["sheet"] = record["sheet"]
    if record.get("markdownLine"):
        ref["markdownLine"] = record["markdownLine"]
    return ref


def build_payload(records, skipped):
    category_counts = count_values(records, "kategori")
    class_counts = count_values(records, "sinif")
    lesson_counts = count_values(records, "ders")
    unit_counts = count_values(records, "unite")
    main_profile_counts = Counter()
    support_profile_counts = Counter()
    for record in records:
        main_profile_counts.update(record["anaProfiller"])
        support_profile_counts.update(record["destekProfiller"])

    profile_counts = main_profile_counts + support_profile_counts
    all_counts = (
        [len(records)]
        + list(category_counts.values())
        + list(class_counts.values())
        + list(lesson_counts.values())
        + list(unit_counts.values())
        + list(profile_counts.values())
    )
    min_count = min(all_counts)
    max_count = max(all_counts)

    nodes = []
    node_ids = {}
    edges = {}

    center_id = "OGRENCI_PROFILI"
    nodes.append(
        {
            "id": center_id,
            "name": "Öğrenci Profili",
            "label": f"Öğrenci\nProfili\n{len(records)} kayıt",
            "group": "MERKEZ",
            "nodeType": "merkez",
            "shape": "circle",
            "size": 48,
            "x": 0,
            "y": 0,
            "fixed": True,
            "mass": 6,
            "color": {"background": "#f8f9fb", "border": PALETTE["MERKEZ"]},
            "font": {"color": "#2f3640", "multi": True, "size": 18, "bold": True},
            "recordCount": len(records),
            "title": f"Öğrenci Profili: {len(records)} kayıt",
        }
    )

    for index, (value, count) in enumerate(sorted(category_counts.items(), key=lambda item: category_sort_key(item[0]))):
        node_id = slug(value, "K")
        node_ids[("kategori", value)] = node_id
        x = -170 if value == ORTAOGRETIM else 170
        nodes.append(
            make_node(
                node_id,
                value,
                "KATEGORI",
                "kategori",
                count,
                min_count,
                max_count,
                "box",
                x=x,
                y=180 + index * 70,
                fixed=True,
                mass=3,
            )
        )
        refs = [record_ref(record) for record in records if record.get("kategori") == value]
        add_edge(edges, center_id, node_id, "eğitim kategorisi", refs)

    sorted_profiles = sorted(profile_counts.items(), key=profile_sort_key)
    profile_angles = {}
    profile_radius = 330
    for index, (profile, count) in enumerate(sorted_profiles):
        angle = -math.pi / 2 + (2 * math.pi * index / max(len(sorted_profiles), 1))
        profile_angles[profile] = angle
        node_id = slug(profile, "P")
        node_ids[("profil", profile)] = node_id
        refs = [
            record_ref(record)
            for record in records
            if profile in record["anaProfiller"] or profile in record["destekProfiller"]
        ]
        nodes.append(
            make_node(
                node_id,
                profile,
                "PROFIL",
                "profil",
                count,
                min_count,
                max_count,
                "dot",
                border=PROFILE_COLORS.get(profile, PALETTE["PROFIL"]),
                x=sector_point(angle, profile_radius)["x"],
                y=sector_point(angle, profile_radius)["y"],
                fixed=True,
                mass=4,
            )
        )
        add_edge(edges, center_id, node_id, "profil alanı", refs)

    for value, count in sorted(class_counts.items(), key=lambda item: class_sort_key(item[0])):
        node_id = slug(value, "S")
        node_ids[("sinif", value)] = node_id
        nodes.append(make_node(node_id, value, "SINIF", "sinif", count, min_count, max_count, "box", mass=2))

    lesson_dominant = {}
    unit_dominant = {}
    for record in records:
        dominant = record["anaProfiller"][0]
        lesson_dominant.setdefault(record["ders"], Counter())
        unit_dominant.setdefault(record["unite"], Counter())
        lesson_dominant[record["ders"]][dominant] += 1
        unit_dominant[record["unite"]][dominant] += 1

    lesson_dominant = {key: counter.most_common(1)[0][0] for key, counter in lesson_dominant.items()}
    unit_dominant = {key: counter.most_common(1)[0][0] for key, counter in unit_dominant.items()}
    lesson_positions = distribute_positions(lesson_counts.keys(), profile_angles, lesson_dominant, 610, spread=0.22)
    unit_positions = distribute_positions(unit_counts.keys(), profile_angles, unit_dominant, 850, spread=0.48)

    for value, count in sorted(lesson_counts.items()):
        node_id = slug(value, "D")
        node_ids[("ders", value)] = node_id
        nodes.append(
            make_node(
                node_id,
                value,
                "DERS",
                "ders",
                count,
                min_count,
                max_count,
                "ellipse",
                **lesson_positions.get(value, {}),
                mass=2,
            )
        )

    for value, count in sorted(unit_counts.items()):
        node_id = slug(value, "U")
        node_ids[("unite", value)] = node_id
        nodes.append(
            make_node(
                node_id,
                value,
                "UNITE_TEMA",
                "unite_tema",
                count,
                min_count,
                max_count,
                "box",
                **unit_positions.get(value, {}),
                mass=1,
            )
        )

    relation_buckets = defaultdict(list)
    for record in records:
        category_id = node_ids[("kategori", record.get("kategori", TEMEL_EGITIM))]
        class_id = node_ids[("sinif", record["sinif"])]
        lesson_id = node_ids[("ders", record["ders"])]
        unit_id = node_ids[("unite", record["unite"])]
        ref = record_ref(record)

        relation_buckets[(category_id, class_id, "kategori sınıfı", False)].append(ref)
        relation_buckets[(class_id, lesson_id, "sınıf dersi", False)].append(ref)
        relation_buckets[(lesson_id, unit_id, "ders ünitesi/teması", False)].append(ref)

        for main_profile in record["anaProfiller"]:
            main_id = node_ids[("profil", main_profile)]
            relation_buckets[(unit_id, main_id, "ana profil", False)].append(ref)

            for support_profile in record["destekProfiller"]:
                support_id = node_ids[("profil", support_profile)]
                relation_buckets[(unit_id, support_id, "destekleyici profil", True)].append(ref)
                relation_buckets[(main_id, support_id, "profil desteği", True)].append(ref)

    for (from_id, to_id, relation, dashed), refs in relation_buckets.items():
        add_edge(edges, from_id, to_id, relation, refs, dashed=dashed)

    node_id_set = {node["id"] for node in nodes}
    missing_edges = [edge for edge in edges.values() if edge["from"] not in node_id_set or edge["to"] not in node_id_set]

    return {
        "title": "Öğrenci Profili Ham Veri Ağ Analizi",
        "description": "Temel Eğitim, Ortaöğretim ve Din Öğretimi sınıf, ders, ünite/tema, ana profil ve destekleyici profil verilerinden üretilmiştir.",
        "source": [
            str(SOURCE_MARKDOWN),
            str(ORTAOGRETIM_SOURCE_DIR),
            str(DIN_OGRETIMI_SOURCE),
            str(SAGLIK_ORTAOGRETIM_SOURCE),
        ],
        "summary": {
            "files": len({record["file"] for record in records}),
            "records": len(records),
            "skippedRows": len(skipped),
            "categories": len(category_counts),
            "recordsByCategory": dict(sorted(category_counts.items(), key=lambda item: category_sort_key(item[0]))),
            "classes": len(class_counts),
            "lessons": len(lesson_counts),
            "units": len(unit_counts),
            "profiles": len(profile_counts),
            "mainProfileAppearances": sum(len(record["anaProfiller"]) for record in records),
            "supportProfileAppearances": sum(len(record["destekProfiller"]) for record in records),
            "missingEdges": len(missing_edges),
        },
        "normalization": PROFILE_ALIASES,
        "groupOrder": GROUP_ORDER,
        "defaultPalette": PALETTE,
        "hcPalette": HC_PALETTE,
        "nodes": nodes,
        "edges": sorted(edges.values(), key=lambda edge: (edge["from"], edge["to"], edge["relation"])),
        "records": records,
        "skippedRows": skipped,
    }


def main():
    records, skipped = read_records()
    ortaogretim_records, ortaogretim_skipped = read_ortaogretim_records()
    din_ogretimi_records, din_ogretimi_skipped = read_din_ogretimi_records()
    ortaogretim_saglik_records, ortaogretim_saglik_skipped = read_ortaogretim_saglik_records()
    records.extend(ortaogretim_records)
    records.extend(din_ogretimi_records)
    records.extend(ortaogretim_saglik_records)
    records = [
        record
        for record in (apply_curriculum_rules(record) for record in records)
        if record["ders"] not in EXCLUDED_LESSONS
    ]
    skipped.extend(ortaogretim_skipped)
    skipped.extend(din_ogretimi_skipped)
    skipped.extend(ortaogretim_saglik_skipped)
    payload = build_payload(records, skipped)
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(payload["summary"], ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
